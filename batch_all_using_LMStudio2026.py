import os
import re
import json
import requests
import numpy as np
from datetime import datetime
from pdf2image import convert_from_path
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment

# ===================== CONFIG =====================

DEBUG = True
API_URL = "http://10.0.0.20:1233/v1/chat/completions"
MODEL_NAME = "google/gemma-3-27b"
POPPLER_PATH = r"C:\Program Files\poppler-25.12.0\Library\bin"

os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

# ===================== LM STUDIO CHECK =====================

def check_lmstudio():
    print("\n🔎 Prüfe LM Studio Verbindung...")
    try:
        r = requests.get("http://10.0.0.20:1233/v1/models", timeout=10)
        r.raise_for_status()
        models = r.json().get("data", [])

        if not models:
            print("❌ Kein Modell geladen!")
            return False

        print("✅ Verbindung ok.")
        for m in models:
            print("   -", m["id"])

        return any(MODEL_NAME in m["id"] for m in models)

    except Exception as e:
        print("❌ Verbindung fehlgeschlagen:", e)
        return False

# ===================== OCR INIT =====================

OCR_ENGINE = PaddleOCR(
    use_textline_orientation=True,
    device="cpu"
)

# ===================== HELPER =====================

def debug(msg):
    if DEBUG:
        print(msg)

def parse_german_float(val):
    if isinstance(val, str):
        val = val.replace(".", "").replace(",", ".")
        try:
            return float(val)
        except:
            return None
    return val if isinstance(val, float) else None

def parse_german_date(val):
    try:
        return datetime.strptime(val, "%d.%m.%Y")
    except:
        return None

# ===================== OCR =====================

def extract_text_with_paddle(pdf_path):

    print("\n=============================")
    print("📄 OCR START:", os.path.basename(pdf_path))
    print("=============================")

    pages = convert_from_path(
        pdf_path,
        dpi=300,
        poppler_path=POPPLER_PATH
    )

    if not pages:
        print("❌ Keine Seiten extrahiert")
        return ""

    full_text = ""

    for page_index, page in enumerate(pages):

        print(f"\n--- Seite {page_index+1} ---")

        img = np.array(page)
        result = OCR_ENGINE.predict(img)

        if not result:
            print("⚠️ OCR Ergebnis leer")
            continue

        print("Result Type:", type(result))
        print("Length:", len(result))

        page_dict = result[0]

        print("Keys im Result:", page_dict.keys())

        texts = page_dict.get("rec_texts", [])
        scores = page_dict.get("rec_scores", [])

        print("Gefundene Texte:", len(texts))

        for txt, score in zip(texts, scores):

            print(f"{score:.3f} | {txt}")

            if score > 0.85:
                full_text += txt.strip() + "\n"

    print("\n=============================")
    print("📄 OCR ENDE")
    print("=============================")

    print("\n--- FINAL TEXT ---")
    print(full_text)
    print("------------------\n")

    return full_text

# ===================== LLM =====================

def send_text_to_llm(ocr_text):

    system_prompt = (
        "Du bist ein extrem präziser deutscher Buchhaltungs-Parser. "
        "Du darfst KEINE Werte erfinden."
    )

    user_prompt = f"""
Extrahiere folgende Rechnungsdaten:

OCR TEXT:
{ocr_text}

Antworte ausschließlich mit gültigem JSON:

{{
"Rechnungsnummer":"",
"Datum":"",
"Bezeichnung":"",
"MwSt-Satz":"",
"MwSt-Betrag":"",
"Gesamtbetrag":"",
"Nettobetrag":"",
"Lieferant":""
}}

Regeln:
- Zahlen mit Komma
- Datum TT.MM.JJJJ
- Wenn nicht vorhanden → NICHT GEFUNDEN
"""

    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.0
    }

    debug("📤 Sende an LLM...")

    response = requests.post(API_URL, json=payload, timeout=300)
    response.raise_for_status()

    content = response.json()["choices"][0]["message"]["content"]

    debug("\n📥 LLM RAW:")
    debug(content)

    return content

# ===================== JSON EXTRACTION =====================

def extract_json(text):
    try:
        return json.loads(text)
    except:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except:
                return {}
    return {}

# ===================== VALIDATION =====================

def validate_amounts(data):

    netto = parse_german_float(data.get("Nettobetrag"))
    mwst = parse_german_float(data.get("MwSt-Betrag"))
    brutto = parse_german_float(data.get("Gesamtbetrag"))

    if netto and mwst and brutto:
        if abs((netto + mwst) - brutto) > 0.02:
            print("⚠️ Beträge inkonsistent!")

# ===================== PROCESS PDF =====================

def process_pdf(pdf_path):

    ocr_text = extract_text_with_paddle(pdf_path)
    llm_answer = send_text_to_llm(ocr_text)
    data = extract_json(llm_answer)

    debug("📦 JSON:", data)

    validate_amounts(data)

    return data, llm_answer

# ===================== EXCEL =====================

def process_folder_to_excel(folder, output_excel):

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Rechnungen"

    headers = [
        "Datei","Rechnungsnummer","Datum","Bezeichnung",
        "MwSt-Satz","MwSt-Betrag","Gesamtbetrag",
        "Nettobetrag","Lieferant","Raw_LLM"
    ]

    for col,h in enumerate(headers,1):
        sheet.cell(row=1,column=col,value=h).font = Font(bold=True)

    row = 2

    for file in sorted(os.listdir(folder)):
        if not file.lower().endswith(".pdf"):
            continue

        path = os.path.join(folder,file)

        try:
            data, raw = process_pdf(path)

            values = [
                file,
                data.get("Rechnungsnummer"),
                parse_german_date(data.get("Datum")),
                data.get("Bezeichnung"),
                data.get("MwSt-Satz"),
                parse_german_float(data.get("MwSt-Betrag")),
                parse_german_float(data.get("Gesamtbetrag")),
                parse_german_float(data.get("Nettobetrag")),
                data.get("Lieferant"),
                raw
            ]

        except Exception as e:
            values = [file] + ["FEHLER"]*8 + [str(e)]

        for col,val in enumerate(values,1):
            sheet.cell(row=row,column=col,value=val)

        row+=1

    wb.save(output_excel)
    print(f"\n✔ Excel gespeichert: {output_excel}")

# ===================== RUN =====================

if __name__ == "__main__":

    if not check_lmstudio():
        print("❌ Gemma nicht erreichbar. Abbruch.")
        exit()

    input_folder = r"C:\Users\surin\Meine Ablage (surinder.ram@gmail.com)\Firma\Belege\2025\Hardware"
    output_excel = r"C:\Users\surin\Meine Ablage (surinder.ram@gmail.com)\Firma\Belege\2025\Hardware\Ergebnis.xlsx"

    process_folder_to_excel(input_folder, output_excel)