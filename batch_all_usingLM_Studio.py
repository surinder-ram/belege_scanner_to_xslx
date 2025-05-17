import os
import csv  # Wird nicht mehr direkt für Output gebraucht, aber lassen wir es für ggf. Debug-CSV
import json
import base64
import io
import requests
import re
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment

from openpyxl.utils.datetime import to_excel

from datetime import datetime

# LM Studio API-Einstellungen
API_URL = "http://10.0.0.20:1233/v1/chat/completions"
MODEL_NAME = "google/gemma-2-27b-it"  # Stelle sicher, dass dies der korrekte Identifier in LM Studio ist
# Oder dein spezifischer Name:
# MODEL_NAME = "gemma-3-27b-it-qat"

# Poppler-Pfad (PDF→Bild)
POPPLER_PATH = r"C:\Program Files\poppler-24.08.0\Library\bin"


# ---- HILFSFUNKTIONEN für Excel ----
def create_excel_styles(workbook):
    """Erstellt und registriert benannte Stile für die Excel-Datei."""
    header_font = Font(bold=True)
    amount_style = NamedStyle(name="amount_style", number_format="#,##0.00")
    date_style = NamedStyle(name="date_style", number_format="DD.MM.YYYY")  # Deutsches Datumsformat

    if "amount_style" not in workbook.style_names:
        workbook.add_named_style(amount_style)
    if "date_style" not in workbook.style_names:
        workbook.add_named_style(date_style)

    return header_font, "amount_style", "date_style"


def parse_german_float(s_val):
    """Konvertiert einen deutschen Zahlenstring (mit Komma) in einen Float."""
    if isinstance(s_val, (int, float)):
        return float(s_val)
    if isinstance(s_val, str):
        s_val = s_val.replace('.', '').replace(',', '.')  # Erst Tausendertrenner entfernen, dann Komma zu Punkt
        try:
            return float(s_val)
        except ValueError:
            return None  # Oder 0.0, je nach gewünschtem Verhalten
    return None


def parse_german_date(s_date):
    """Konvertiert einen deutschen Datumsstring (TT.MM.JJJJ) in ein datetime-Objekt."""
    if isinstance(s_date, str):
        try:
            return datetime.strptime(s_date, "%d.%m.%Y")
        except ValueError:
            return None
    return None


# ---- ENDE HILFSFUNKTIONEN für Excel ----

def extract_data_from_pdf_image_directly(pdf_path, page_number=0, dpi=300):
    """
    Lädt eine PDF-Seite als Bild, encodiert sie und schickt sie zusammen
    mit einem detaillierten Extraktions-Prompt an das Vision-LLM,
    um direkt ein JSON-Objekt mit den Rechnungsdaten zu erhalten.
    """
    print(f"    Versuche PDF-Seite {page_number} von '{os.path.basename(pdf_path)}' zu laden (DPI: {dpi})...")
    try:
        pages = convert_from_path(
            pdf_path,
            dpi=dpi,
            poppler_path=POPPLER_PATH,
            first_page=page_number + 1,
            last_page=page_number + 1
        )
    except Exception as e:
        print(f"    Fehler beim Konvertieren von PDF zu Bild für {pdf_path}: {e}")
        raise

    if not pages:
        print(f"    Fehler: PDF {pdf_path} hat keine Seite {page_number} oder konnte nicht gelesen werden.")
        raise ValueError(f"PDF {pdf_path} hat keine Seite {page_number} oder konnte nicht gelesen werden.")
    pil_img = pages[0]
    print(f"    Bild von '{os.path.basename(pdf_path)}' geladen (Größe: {pil_img.width}x{pil_img.height}).")

    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    image_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    print(f"    Bild für '{os.path.basename(pdf_path)}' base64-encodiert.")

    system_prompt_content = (
        "Du bist ein Experte für die Analyse von Rechnungsbildern. "
        "Deine Aufgabe ist es, die relevanten Informationen direkt aus dem übergebenen Bild zu extrahieren und "
        "ausschließlich im geforderten JSON-Format zurückzugeben. "
        "Verwende für Dezimalzahlen das deutsche Komma (z.B. 19,00). "
        "Gib bei allen Beträgen und Sätzen ausschließlich die Zahl zurück, OHNE Euro-Zeichen oder Prozent-Zeichen. "
        "Wenn ein Wert nicht gefunden werden kann, schreibe den String \"NICHT GEFUNDEN\" für das entsprechende Feld."
    )
    user_prompt_content = (
        "Analysiere das folgende Rechnungsbild und extrahiere die unten aufgeführten Felder. "
        "Gib deine Antwort ausschließlich als JSON-Objekt zurück, wie im Beispiel gezeigt.\n\n"
        "Zu extrahierende Felder:\n"
        "- Rechnungsnummer\n"
        "- Datum (Verwende für das Datum das Format TT.MM.JJJJ, z.B. 05.12.2024)\n"
        "- Bezeichnung (Falls es eine Restaurant-Rechnung ist, nimm den Restaurant-Namen. Für Supermärkte 'Einkauf' oder den Supermarkt-Namen, falls klar ersichtlich.)\n"
        "- MwSt-Satz (in Prozent, nur Zahl, KEIN %, KEIN Punkt, sondern Komma! z.B. '19,00'. Wenn mehrere Sätze vorhanden sind, nimm den dominanten oder wichtigsten.)\n"
        "- MwSt-Betrag (nur Zahl, KEIN €. Wenn mehrere MwSt-Beträge vorhanden sind, summiere sie oder nimm den zur Haupt-MwSt gehörenden Betrag.)\n"
        "- Gesamtbetrag (nur Zahl, KEIN €, inkl. MwSt)\n"
        "- Nettobetrag (nur Zahl, KEIN €, ohne MwSt. Berechne diesen ggf. aus Gesamtbetrag und MwSt-Betrag, falls nicht explizit genannt.)\n"
        "- Lieferant (Name des Unternehmens, das die Rechnung ausgestellt hat)\n\n"
        "Beispiel für das JSON-Antwortformat:\n"
        "{\n"
        "  \"Rechnungsnummer\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"Datum\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"Bezeichnung\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"MwSt-Satz\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"MwSt-Betrag\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"Gesamtbetrag\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"Nettobetrag\": \"WERT_ODER_NICHT_GEFUNDEN\",\n"
        "  \"Lieferant\": \"WERT_ODER_NICHT_GEFUNDEN\"\n"
        "}\n\n"
        "Achtung: Schreibe niemals einen Punkt als Dezimaltrennzeichen, sondern IMMER ein Komma!\n"
        "Stelle sicher, dass alle angeforderten Felder im JSON enthalten sind, auch wenn der Wert \"NICHT GEFUNDEN\" ist."
    )
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_prompt_content},
            {"role": "user", "content": [
                {"type": "text", "text": user_prompt_content},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}}
            ]}
        ],
        "temperature": 0.1,
        "max_tokens": 2048,
        "stream": False
    }
    headers = {"Content-Type": "application/json"}

    print(f"    Sende Anfrage an LLM für '{os.path.basename(pdf_path)}'...")
    try:
        response = requests.post(API_URL, json=payload, headers=headers, timeout=300)  # Timeout evtl. weiter erhöhen
        response.raise_for_status()
        data = response.json()
        print(f"    Antwort vom LLM für '{os.path.basename(pdf_path)}' erhalten.")
        if "choices" in data and data["choices"] and "message" in data["choices"][0] and "content" in \
                data["choices"][0]["message"]:
            return data["choices"][0]["message"]["content"]
        else:
            print(f"    Unerwartete LLM-Antwortstruktur: {json.dumps(data)}")
            return "{}"
    except requests.exceptions.RequestException as e:
        print(f"    API-Anfragefehler für '{os.path.basename(pdf_path)}': {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"    API Response Text: {e.response.text}")
        raise RuntimeError(f"API-Anfragefehler: {e}")
    except Exception as e:
        print(f"    Allgemeiner Fehler in extract_data_from_pdf_image_directly für '{os.path.basename(pdf_path)}': {e}")
        return "{}"


def extract_json_from_response(response_text):
    if not isinstance(response_text, str):
        print(f"    Warnung: extract_json_from_response erwartet String, bekam {type(response_text)}")
        return {}
    text_to_parse = response_text.strip()
    match_markdown = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text_to_parse, re.DOTALL)
    if match_markdown:
        json_str = match_markdown.group(1)
    else:
        match_json = re.search(r"\{[\s\S]*\}", text_to_parse)
        if match_json:
            json_str = match_json.group(0)
        else:
            print("    Kein JSON-Objekt im LLM-Antworttext gefunden.")
            # print(f"    Antworttext war: >>>{text_to_parse}<<<") # Optional für Debugging
            return {}
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"    JSON-Parsing-Fehler: {e}")
        # print(f"    Fehlerhafter JSON-String: >>>{json_str}<<<") # Optional für Debugging
        return {}


def process_pdf_and_get_data(pdf_path):
    """Verarbeitet eine einzelne PDF und gibt die extrahierten Daten und die rohe LLM-Antwort zurück."""
    print(f"  Verarbeite Datei: {os.path.basename(pdf_path)}")
    llm_result_str = extract_data_from_pdf_image_directly(pdf_path, dpi=300)
    print(f"    Rohe LLM-Antwort für {os.path.basename(pdf_path)}:\n{llm_result_str[:200]}...")  # Gekürzte Ausgabe

    result_json = extract_json_from_response(llm_result_str)
    if not result_json:
        print(f"    Konnte keine valide JSON-Antwort für {os.path.basename(pdf_path)} parsen.")
        return None, llm_result_str  # Daten sind None, wenn JSON-Parsing fehlschlägt

    print(f"    Extrahiertes JSON für {os.path.basename(pdf_path)}:")
    print(f"      Rechnungsnummer: {result_json.get('Rechnungsnummer')}")
    print(f"      Datum: {result_json.get('Datum')}")
    print(f"      Gesamtbetrag: {result_json.get('Gesamtbetrag')}")
    print(f"      Lieferant: {result_json.get('Lieferant')}")
    # Weitere Felder bei Bedarf ausgeben

    return result_json, llm_result_str


def process_folder_to_excel(folder_path, output_excel_path):
    """
    Verarbeitet alle PDFs in einem Ordner und schreibt die Ergebnisse in eine Excel-Datei.
    """
    print(f"\nStarte Verarbeitung für Ordner: {folder_path}")
    print(f"Ausgabedatei wird: {output_excel_path}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rechnungsdaten"

    header_font, amount_style_name, date_style_name = create_excel_styles(workbook)

    headers = [
        "Datei", "Rechnungsnummer", "Datum", "Bezeichnung", "MwSt-Satz",
        "MwSt-Betrag", "Gesamtbetrag", "Nettobetrag", "Lieferant", "Rohe_LLM_Antwort"
    ]
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    for dateiname in sorted(os.listdir(folder_path)):  # Sortiert für konsistente Reihenfolge
        if not dateiname.lower().endswith(".pdf"):
            continue

        pdf_full_path = os.path.join(folder_path, dateiname)

        try:
            extracted_data, raw_llm_response = process_pdf_and_get_data(pdf_full_path)

            # Daten für Excel vorbereiten
            row_values = [dateiname]
            if extracted_data:
                row_values.extend([
                    extracted_data.get("Rechnungsnummer", "NICHT GEFUNDEN"),
                    parse_german_date(extracted_data.get("Datum")),  # Konvertiere zu datetime
                    extracted_data.get("Bezeichnung", "NICHT GEFUNDEN"),
                    extracted_data.get("MwSt-Satz", "NICHT GEFUNDEN"),  # Bleibt als String, da es '%' sein könnte
                    parse_german_float(extracted_data.get("MwSt-Betrag")),
                    parse_german_float(extracted_data.get("Gesamtbetrag")),
                    parse_german_float(extracted_data.get("Nettobetrag")),
                    extracted_data.get("Lieferant", "NICHT GEFUNDEN"),
                    raw_llm_response
                ])
            else:  # Fehlerfall oder kein JSON
                row_values.extend(["FEHLER"] * 8 + [raw_llm_response if raw_llm_response else "Keine LLM Antwort"])

            for col_num, value in enumerate(row_values, 1):
                cell = sheet.cell(row=current_row, column=col_num, value=value)
                # Stile anwenden
                if headers[col_num - 1] == "Datum" and isinstance(value, datetime):
                    cell.style = date_style_name
                elif headers[col_num - 1] in ["MwSt-Betrag", "Gesamtbetrag", "Nettobetrag"] and isinstance(value,
                                                                                                           float):
                    cell.style = amount_style_name

            current_row += 1

        except Exception as e:
            print(f"  Schwerwiegender Fehler bei der Hauptverarbeitung von {dateiname}: {e}")
            # Fehlerzeile in Excel schreiben
            error_row = [dateiname] + ["FEHLER"] * 8 + [str(e)]
            for col_num, value in enumerate(error_row, 1):
                sheet.cell(row=current_row, column=col_num, value=value)
            current_row += 1
        print("-" * 30)  # Trennlinie zwischen Dateien

    # Spaltenbreiten anpassen (einfacher Ansatz)
    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter  # Spaltenbuchstabe
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if column_letter == 'J':  # Rohe LLM Antwort
            adjusted_width = min(adjusted_width, 50)  # Begrenzen, damit es nicht zu breit wird
        sheet.column_dimensions[column_letter].width = adjusted_width

    try:
        workbook.save(output_excel_path)
        print(f"\nErfolgreich gespeichert: {output_excel_path}")
    except Exception as e:
        print(f"\nFehler beim Speichern der Excel-Datei {output_excel_path}: {e}")


def einzelfile_verarbeiten():
    """Verarbeitet eine einzelne, fest codierte PDF-Datei."""
    print("=== EINZELDATEI-VERARBEITUNG START ===")
    # Passe den Pfad zu deiner Testdatei an
    test_file = r"C:\Users\surin\Meine Ablage (surinder.ram@gmail.com)\Firma\Belege\2024\Bewirtung\208_Hofer_Bewirtung_12_10_2024.pdf"

    if not os.path.exists(test_file):
        print(f"FEHLER: Testdatei nicht gefunden: {test_file}")
        return

    try:
        extracted_data, raw_llm_response = process_pdf_and_get_data(test_file)

        if extracted_data:
            print("\n--- Finale extrahierte und geparste Daten für Einzelfile ---")
            print(f"  Datei: {os.path.basename(test_file)}")
            print(f"  Rechnungsnummer: {extracted_data.get('Rechnungsnummer')}")
            datum_obj = parse_german_date(extracted_data.get('Datum'))
            print(f"  Datum: {datum_obj.strftime('%d.%m.%Y') if datum_obj else 'NICHT GEFUNDEN'}")
            print(f"  Bezeichnung: {extracted_data.get('Bezeichnung')}")
            print(f"  MwSt-Satz: {extracted_data.get('MwSt-Satz')}")
            print(f"  MwSt-Betrag: {parse_german_float(extracted_data.get('MwSt-Betrag'))}")
            print(f"  Gesamtbetrag: {parse_german_float(extracted_data.get('Gesamtbetrag'))}")
            print(f"  Nettobetrag: {parse_german_float(extracted_data.get('Nettobetrag'))}")
            print(f"  Lieferant: {extracted_data.get('Lieferant')}")
            # print(f"  Rohe LLM Antwort (gekürzt): {raw_llm_response[:300]}...") # Falls benötigt
        else:
            print(f"\nKeine validen Daten für {os.path.basename(test_file)} extrahiert.")
            print(f"  Rohe LLM Antwort war: {raw_llm_response}")

    except Exception as e:
        print(f"Fehler bei der Verarbeitung der Einzelfile {os.path.basename(test_file)}: {e}")
    print("=== EINZELDATEI-VERARBEITUNG ENDE ===")


def batch_verarbeiten():
    """Verarbeitet alle Unterordner im Hauptverzeichnis."""
    print("\n=== BATCH-VERARBEITUNG START ===")
    main_belege_folder = r"C:\Users\surin\Meine Ablage (surinder.ram@gmail.com)\Firma\Belege\2024"

    if not os.path.exists(main_belege_folder):
        print(f"FEHLER: Hauptordner nicht gefunden: {main_belege_folder}")
        return

    for item_name in os.listdir(main_belege_folder):
        item_path = os.path.join(main_belege_folder, item_name)
        if os.path.isdir(item_path):
            print(f"\nBearbeite Unterordner: {item_name}")
            # Excel-Datei im Hauptordner speichern, benannt nach dem Unterordner
            excel_filename = f"Extraktion_{item_name.replace(' ', '_')}.xlsx"
            output_excel_file_path = os.path.join(main_belege_folder, excel_filename)
            process_folder_to_excel(item_path, output_excel_file_path)
    print("=== BATCH-VERARBEITUNG ENDE ===")


if __name__ == "__main__":
    # Wähle eine der beiden Optionen aus, indem du die andere auskommentierst:

    #einzelfile_verarbeiten()
    batch_verarbeiten()